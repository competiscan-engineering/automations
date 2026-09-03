"""
report_lib.py
─────────────────────────────────────────────────────────────────────────────
Shared machinery for the Competiscan recurring-report pipelines.

Reports #3…#200 should be thin config files that import from here. This module
holds everything that is NOT specific to one report:

  Bedrock/Claude   : bedrock_client(), MODEL_ARN, call_claude()
  LLM plumbing     : extract_json(), pick_ids(), cap_text(), clean_cell()
  Concurrency      : run_parallel()          (asyncio thread pool, order-preserving)
  Deck output      : save_pptx()             (base64 | filepath | locked-file fallback)
  Excel output     : write_workbook()        (generic openpyxl writer w/ hyperlinks)
  Email delivery   : send_email()            (AWS SES, raw MIME w/ attachments)
                     notify_report_ready()   (standard subject/body wrapper — use this one)
  MCP tool loading : load_tool()             (sys.path shim + fastmcp unwrap)

Claude is reached through AWS Bedrock (same inference-profile ARN as the chat
servers; boto3 default credential chain — no ANTHROPIC_API_KEY). Factored,
verbatim in behavior, from report_MonthlyBankingMerger.py — the hard-won
defensive edge cases are preserved.
─────────────────────────────────────────────────────────────────────────────
"""

import os
import re
import sys
import json
import html as html_lib
import base64
import asyncio
import importlib
import threading
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Optional
from urllib.parse import urljoin

import requests
import boto3
from botocore.config import Config

# Repo root — this module lives in pipelines/, so the root is one level up.
# mcp/ (the tool modules) is a sibling of pipelines/, resolved off this.
ROOT = Path(__file__).resolve().parent.parent

# Load variables from the project-root .env into the environment, if present.
# Optional — no-op if python-dotenv isn't installed (pip install python-dotenv).
# Every report_*.py pipeline imports this module, so this is the one place
# that guarantees PPT_BUILDER_LOGIN/PASSWORD (and anything else in .env) are
# in os.environ before anything tries to read them.
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ModuleNotFoundError:
    pass

# ── Bedrock ──────────────────────────────────────────────────────────────────
# Same inference-profile ARN the chat servers use; override via env for testing.
MODEL_ARN = os.environ.get(
    "BEDROCK_MODEL_ARN",
    "arn:aws:bedrock:us-east-2:078398740737:application-inference-profile/myhse60a8wao",
)
# Dollars per token. Bedrock bills per million, so the division is written out rather
# than folded away — a price change is then a one-character edit against the published
# rate card instead of arithmetic done in your head.
_MODEL_IN_PRICE  = 3.00  / 1_000_000
_MODEL_OUT_PRICE = 15.00 / 1_000_000


def _region_from_arn(arn: str) -> str:
    parts = arn.split(":")
    return parts[3] if len(parts) > 3 and parts[3] else "us-east-2"


_BEDROCK_CONFIG = Config(
    region_name=_region_from_arn(MODEL_ARN),
    retries={"max_attempts": 3, "mode": "adaptive"},
    read_timeout=120,
    connect_timeout=10,
)


def bedrock_client():
    """A bedrock-runtime client with the shared retry/timeout config."""
    return boto3.client("bedrock-runtime", config=_BEDROCK_CONFIG)


# Set RS_LLM_TRACE=1 to have every model call report its token usage and what it
# actually said. OFF by default and deliberately so: a scheduled pipeline writing this
# into production logs would be noise, and the responses can carry client detail.
# Pipelines Studio switches it on for runs it starts itself, and prints the result to
# the terminal the Studio was launched from rather than into the researcher's output
# panel — it is engineering telemetry, not something a researcher asked to see.
LLM_TRACE = os.environ.get("RS_LLM_TRACE") == "1"
LLM_TRACE_PREFIX = "[LLM]"

# What this PROCESS has spent. The write-ups and the picks run on a thread pool, so the
# read-modify-write needs a lock or the count quietly comes out short.
_LLM_SPEND = {"calls": 0, "in": 0, "out": 0, "usd": 0.0}
_LLM_LOCK = threading.Lock()


def llm_spend() -> dict:
    """A snapshot of what this process has spent on the model."""
    with _LLM_LOCK:
        return dict(_LLM_SPEND)


def llm_cost(in_tokens: int, out_tokens: int) -> float:
    return (in_tokens or 0) * _MODEL_IN_PRICE + (out_tokens or 0) * _MODEL_OUT_PRICE


def _trace_llm(payload: dict, system: str, prompt: str, text: str) -> None:
    """One block per model call, on stdout, every line prefixed so a reader can pick
    it back out of an interleaved stream."""
    usage = payload.get("usage") or {}
    tin = int(usage.get("input_tokens") or 0)
    tout = int(usage.get("output_tokens") or 0)
    cache_r = usage.get("cache_read_input_tokens") or 0
    cache_w = usage.get("cache_creation_input_tokens") or 0
    cost = llm_cost(tin, tout)
    with _LLM_LOCK:
        _LLM_SPEND["calls"] += 1
        _LLM_SPEND["in"] += tin
        _LLM_SPEND["out"] += tout
        _LLM_SPEND["usd"] += cost
        so_far = dict(_LLM_SPEND)

    # The prompts are built from a template, so the first line of the system prompt is
    # the cheapest reliable way to say which of the three calls this was.
    kind = (system or "").strip().splitlines()[0][:70] if system else "?"

    out = [
        # ASCII only: this is read on a cp1252 console as often as not, and a
        # box-drawing character would take the whole trace out with a UnicodeEncodeError.
        #
        # cost= is also the field Pipelines Studio adds up to report a whole run's
        # spend, so it stays machine-readable: six decimal places, no thousands
        # separators, one per line.
        f"{LLM_TRACE_PREFIX} -- {MODEL_ARN.rsplit('/', 1)[-1]} "
        f"in={tin} out={tout} stop={payload.get('stop_reason')} "
        f"cost=${cost:.6f}"
        + (f" cache_read={cache_r}" if cache_r else "")
        + (f" cache_write={cache_w}" if cache_w else ""),
        f"{LLM_TRACE_PREFIX}    call: {kind}",
        f"{LLM_TRACE_PREFIX}    prompt: {len(prompt or '')} chars",
    ]
    for line in (text or "").splitlines() or [""]:
        out.append(f"{LLM_TRACE_PREFIX}    > {line}")
    out.append(
        f"{LLM_TRACE_PREFIX}    this process so far: {so_far['calls']} call(s)  "
        f"in={so_far['in']:,}  out={so_far['out']:,}  ${so_far['usd']:.4f}")
    print("\n".join(out), flush=True)


def call_claude(system: str, prompt: str, max_tokens: int = 4000) -> str:
    """Single non-streaming invoke_model call. Warns on a max_tokens stop and
    returns the response text."""
    body = {
        "anthropic_version": "bedrock-2023-05-31",
        "system": system,
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": max_tokens,
    }
    resp = bedrock_client().invoke_model(modelId=MODEL_ARN, body=json.dumps(body))
    payload = json.loads(resp["body"].read())
    if payload.get("stop_reason") == "max_tokens":
        print("   ! Warning: response hit max_tokens — output may be truncated.")
    text = next(
        (b.get("text", "") for b in payload.get("content", []) if b.get("type") == "text"),
        "",
    )
    if LLM_TRACE:
        try:
            _trace_llm(payload, system, prompt, text)
        except Exception:      # telemetry must never take a run down
            pass
    return text


# ── LLM output plumbing ──────────────────────────────────────────────────────
def extract_json(text: str) -> dict:
    """Parse a model response into a dict, tolerating code fences / preamble by
    slicing the outermost {…}."""
    text = (text or "").strip()
    if text.startswith("```"):
        text = re.sub(r"^```[a-zA-Z]*\s*", "", text)
        text = re.sub(r"\s*```$", "", text).strip()
    start, end = text.find("{"), text.rfind("}")
    if start != -1 and end > start:
        text = text[start:end + 1]
    return json.loads(text)


def pick_ids(chosen, records: list[dict], want: int,
             max_ids: int = 5, id_key: str = "entry_id", exclude=None) -> list[str]:
    """Return up to `want` unique, existing IDs (clamped to `max_ids`).

    Chosen IDs that exist in `records` come first (deduped, order preserved);
    if the model returned fewer than `want`, top up from the most-recent records.
    IDs in `exclude` are never picked — use this to keep an entry_id from being
    featured on more than one slide across sequential calls (pass a set that
    accumulates the IDs picked so far and update it with this call's result).
    """
    exclude = exclude or set()
    available = [r[id_key] for r in records if r.get(id_key) and r[id_key] not in exclude]
    avail_set = set(available)
    seen: set = set()
    picked: list = []
    for c in list(chosen or []) + available:  # chosen first, then top-up
        if c in avail_set and c not in seen:
            seen.add(c)
            picked.append(c)
            if len(picked) >= want:
                break
    return picked[:max_ids]


def chunk_ids(ids: list[str], size: int = 5) -> list[list[str]]:
    """Split a list of ids into consecutive chunks of at most `size` (for slide
    types that cap how many entry_ids fit on one slide)."""
    return [ids[i:i + size] for i in range(0, len(ids), size)]


def cap_text(text: str, limit: int) -> str:
    """Word-boundary trim to `limit` chars, appending an ellipsis if trimmed."""
    text = (text or "").strip()
    if len(text) <= limit:
        return text
    cut = text[:limit - 1]
    if " " in cut:
        cut = cut.rsplit(" ", 1)[0]
    return cut.rstrip(" .,;:") + "…"


def clean_cell(val: Any) -> str:
    """Strip newlines / surrounding whitespace from a value for CSV-ish fields."""
    return " ".join(str(val if val is not None else "").split())


def as_text(value) -> str:
    """Coerce an LLM JSON field to a printable string, defensively — a prompt
    can ask for a single string and still get a list/dict back sometimes.
    Console-debug-output-safe: never let the field's shape crash the caller."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, list):
        return "\n".join(as_text(v) for v in value)
    if isinstance(value, dict):
        return "\n".join(f"{k}: {as_text(v)}" for k, v in value.items())
    return str(value).strip()


def fit_text(text: str, limit: int) -> str:
    """Trim to whole sentences that fit under `limit` chars — NO ellipsis, never
    cut mid-sentence. Returns as-is if already short enough. Only if a single
    leading sentence already exceeds the limit does it hard-cut on a word
    boundary (still no ellipsis)."""
    text = " ".join((text or "").split())
    if len(text) <= limit:
        return text
    out = ""
    for part in re.split(r"(?<=[.!?])\s+", text):
        candidate = f"{out} {part}".strip() if out else part
        if len(candidate) <= limit:
            out = candidate
        else:
            break
    if out:
        return out
    cut = text[:limit]
    return (cut.rsplit(" ", 1)[0] if " " in cut else cut).rstrip(" .,;:")


def cap_words(text: str, max_words: int) -> str:
    """Trim to whole sentences that fit within `max_words` words — NO ellipsis,
    never cut mid-sentence. Same shape as fit_text but counting words instead
    of characters (for slide fields sized by word count rather than char
    count). Only if a single leading sentence already exceeds the cap does it
    hard-cut on a word boundary (still no ellipsis)."""
    text = " ".join((text or "").split())
    words = text.split(" ") if text else []
    if len(words) <= max_words:
        return text
    out = ""
    for part in re.split(r"(?<=[.!?])\s+", text):
        candidate = f"{out} {part}".strip() if out else part
        if len(candidate.split(" ")) <= max_words:
            out = candidate
        else:
            break
    if out:
        return out
    return " ".join(words[:max_words]).rstrip(" .,;:")


# ── Concurrency ──────────────────────────────────────────────────────────────
def run_parallel(jobs: list[Callable[[], Any]]) -> list:
    """Run blocking callables concurrently on a thread pool (boto3/requests-safe)
    and return results in input order. A job that raises is surfaced as
    {"error": str(exc)} rather than crashing the batch."""
    if not jobs:
        return []

    async def _run():
        loop = asyncio.get_event_loop()

        async def _one(fn):
            try:
                return await loop.run_in_executor(None, fn)
            except Exception as exc:  # noqa: BLE001 — surface, don't crash the batch
                return {"error": str(exc)}

        return await asyncio.gather(*[_one(j) for j in jobs])

    return asyncio.run(_run())


# ── Deck output ──────────────────────────────────────────────────────────────
def save_pptx(result: Any, out_path) -> Path:
    """Persist a build_deck_* result.

    Handles: a dict carrying an "error"; a "pptx_base64" payload (decoded and
    written, with a timestamped-name fallback if the target file is locked/open
    in PowerPoint); or, if base64 is absent, the builder's own saved "filepath"
    (reported as-is, per standing preference). Returns the path actually written.
    """
    out_path = Path(out_path)
    if not isinstance(result, dict):
        raise RuntimeError(f"PPT builder returned unexpected value: {result!r}")
    if "error" in result:
        raise RuntimeError(f"PPT builder error: {result['error']}")

    b64 = result.get("pptx_base64")
    if b64:
        raw = base64.b64decode(b64)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            out_path.write_bytes(raw)
            return out_path
        except PermissionError:
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            alt = out_path.with_name(f"{out_path.stem}_{stamp}{out_path.suffix}")
            alt.write_bytes(raw)
            return alt

    # No base64, but the builder wrote a file itself — report that path.
    fp = result.get("filepath")
    if fp and Path(fp).is_file():
        return Path(fp)

    raise RuntimeError(f"PPT builder returned no pptx_base64 and no filepath: {result}")


# ── Excel output ─────────────────────────────────────────────────────────────
_HYPERLINK_COLOR = "FF0000FF"
_DEFAULT_WIDTHS = {  # per-header column widths; sensible fallback otherwise
    "Headline": 42, "Product": 30, "Primary Company": 24, "Additional Companies": 24,
    "Primary Sector": 16, "Primary Category": 20, "Primary Sub Category": 22,
    "Primary Sub Sub Category": 24,
    "PDF Content": 14, "EntryID": 13, "Quarter": 10, "Media Channel": 18,
    "Market": 16, "State/Province": 16, "Age": 10, "Income": 14,
    "Mailing Type": 20, "Publication": 20, "Network Name": 18,
    "Social Media Ad Type": 20, "Pre-Screen": 12,
    "Mortgage & Loan - Application Type": 26,
}
_TOKEN = re.compile(r"\{(\w+)\}")


def _fill(template: str, row: dict) -> str:
    return _TOKEN.sub(lambda m: str(row.get(m.group(1), "") or ""), template)


def _tokens_present(template: str, row: dict) -> bool:
    return all(row.get(k) not in (None, "") for k in _TOKEN.findall(template))


def _xl_escape(s: str) -> str:
    return s.replace('"', '""')


def write_workbook(path, sheets: list[dict]) -> Path:
    """Generic multi-sheet openpyxl writer.

    Each sheet dict:
        name       : worksheet title
        filter_row : full A1 filter string (bold, row 1)
        headers    : exact column names (bold, row 2)
        rows       : list of {header: value} dicts (row 3+)
        hyperlinks : optional {column: (url_template, display_template)} where
                     {token} is filled from the row dict (e.g. {pid}, {entry_id});
                     rendered as =HYPERLINK(...) with hyperlink styling. If a
                     required token is missing, the display text is written plain.

    Data starts at row 3; panes are frozen at A3. `Additional Companies` defaults
    to "N/A". One worksheet per spec, in the order given.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    link_font = Font(color=_HYPERLINK_COLOR, underline="single")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    for spec in sheets:
        ws = wb.create_sheet(title=str(spec["name"])[:31])  # Excel 31-char limit
        headers = list(spec["headers"])
        hyperlinks = spec.get("hyperlinks", {}) or {}
        header_fills = spec.get("header_fills", {}) or {}  # {column_name: RGB hex}

        ws.cell(row=1, column=1, value=spec.get("filter_row", "")).font = bold
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=c, value=h)
            cell.font = bold
            if h in header_fills:
                cell.fill = PatternFill(fill_type="solid", fgColor=header_fills[h])
            ws.column_dimensions[get_column_letter(c)].width = _DEFAULT_WIDTHS.get(h, 18)

        for r, row in enumerate(spec.get("rows", []), start=3):
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=r, column=c)
                if h in hyperlinks:
                    url_tmpl, disp_tmpl = hyperlinks[h]
                    disp = _fill(disp_tmpl, row)
                    if _tokens_present(url_tmpl, row):
                        url = _fill(url_tmpl, row)
                        cell.value = f'=HYPERLINK("{_xl_escape(url)}","{_xl_escape(disp)}")'
                        cell.font = link_font
                    else:
                        cell.value = disp  # no id token → plain text, flagged upstream
                    continue
                val = row.get(h, "")
                if h == "Additional Companies" and not val:
                    val = "N/A"
                cell.value = "" if val is None else val

        ws.freeze_panes = "A3"

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
        return path
    except PermissionError:
        # File is open (locked) in Excel — save a timestamped sibling instead.
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = path.with_name(f"{path.stem}_{stamp}{path.suffix}")
        wb.save(alt)
        return alt


# ── Email delivery ────────────────────────────────────────────────────────────
# AWS SES — same boto3 default credential chain as Bedrock, no separate SMTP
# host/user/password needed. Region is independently overridable since SES
# sending-identity verification is per-region and may not live in the same
# region as the Bedrock inference profile.
SES_REGION       = os.environ.get("SES_REGION", "us-east-1")
DEFAULT_SENDER   = "pipelines@ai.competiscan.com"
DEFAULT_RECIPIENT = "hgquijano@competiscan.com"


def _ses_client():
    return boto3.client("sesv2", region_name=SES_REGION)


def send_email(
    attachment_paths,
    to_addr: str = DEFAULT_RECIPIENT,
    subject: str = "Competiscan report",
    body: str = "Attached is the latest report.",
    from_addr: str = DEFAULT_SENDER,
) -> dict:
    """Email one or more report files (PPTX/XLSX/etc.) via AWS SES.

    REQUIRES `from_addr` (or its whole domain) to already be a VERIFIED SES
    sending identity in the target AWS account/region — this function cannot
    create that verification itself; SES will simply reject the send until
    that's done. Failures are returned as a dict, never raised, so a missing
    prerequisite here can't crash a pipeline run that already has its deck/
    Excel safely written to disk — check `result["status"]` and log/print a
    warning rather than treating this as fatal.

    Args:
        attachment_paths : one path (str/Path) or a list of them, attached
                            as-is, under their own filename.
        to_addr           : recipient email address.
        subject           : email subject line.
        body              : plain-text email body.
        from_addr         : sender address — must be an SES-verified identity.

    Returns:
        {"status": "sent", "message_id": ...} on success, or
        {"status": "error", "error": str(...)} on failure (missing attachment,
        unverified identity, throttling, etc. — all surfaced the same way).
    """
    from email.mime.application import MIMEApplication
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    paths = [attachment_paths] if isinstance(attachment_paths, (str, Path)) else list(attachment_paths)
    paths = [Path(p) for p in paths]

    missing = [str(p) for p in paths if not p.is_file()]
    if missing:
        return {"status": "error", "error": f"attachment(s) not found: {missing}"}

    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"] = from_addr
    msg["To"] = to_addr
    msg.attach(MIMEText(body, "plain"))

    for p in paths:
        part = MIMEApplication(p.read_bytes(), Name=p.name)
        part["Content-Disposition"] = f'attachment; filename="{p.name}"'
        msg.attach(part)

    try:
        resp = _ses_client().send_email(
            FromEmailAddress=from_addr,
            Destination={"ToAddresses": [to_addr]},
            Content={"Raw": {"Data": msg.as_bytes()}},  # SESv2 Raw.Data is a blob — bytes, not str
        )
        return {"status": "sent", "message_id": resp.get("MessageId")}
    except Exception as exc:  # noqa: BLE001 — surface, don't crash the pipeline
        return {"status": "error", "error": str(exc)}


def notify_report_ready(
    report_name: str,
    period_label: str,
    attachment_paths,
    to_addr: str = DEFAULT_RECIPIENT,
    from_addr: str = DEFAULT_SENDER,
) -> dict:
    """Email a finished report's deliverables to a reviewer, in the standard
    subject/body used across every recurring-report pipeline. Thin wrapper
    around send_email() — call THIS (not send_email directly) from any
    report_*.py so every pipeline's end-of-run notification reads the same
    way, with the same review reminder.

    Args:
        report_name      : human-readable report name, e.g. "SupplyHouse.com
                            Competitor Ads" or "Harborstone Weekly Update".
        period_label      : the period covered, e.g. "June 2026" or
                            "July 14th, 2026".
        attachment_paths  : one path or a list of paths (PPTX/XLSX/etc.).
        to_addr           : reviewer's email address.
        from_addr         : sender — must be an SES-verified identity.

    Returns:
        Same shape as send_email(): {"status": "sent"/"error", ...}.
    """
    paths = [attachment_paths] if isinstance(attachment_paths, (str, Path)) else list(attachment_paths)
    paths = [Path(p) for p in paths]
    file_list = "\n".join(f"  - {p.name}" for p in paths) or "  (no files attached)"

    subject = f"{report_name} — {period_label} is ready for review"
    body = (
        f"Hey {to_addr},\n\n"
        f"AI pipelines just finished the {report_name} trend report for {period_label}. "
        f"Deliverables are attached:\n\n"
        f"{file_list}\n\n"
        f"This was generated autonomously by AI agents — please review it before sending "
        f"to the client.\n\n"
        f"— Competiscan Pipelines, powered by the Brain"
    )
    return send_email(paths, to_addr=to_addr, subject=subject, body=body, from_addr=from_addr)


# ── PPT Builder auth (ALB + Cognito) ─────────────────────────────────────────
# csresearchhub.com (the PPT Builder API every report_*.py pipeline posts decks
# to, via mcp_pptbuilder.build_deck_*) sits behind an AWS ALB with a Cognito
# user-pool authorizer. Unauthenticated requests get redirected to Cognito's
# hosted login UI; on success the ALB itself sets AWSELBAuthSessionCookie-0/1
# cookies and, on every subsequent request, injects the x-amzn-oidc-* headers
# Flask trusts. This lives here (not in mcp_pptbuilder.py) so every pipeline
# that talks to PPT Builder shares one login implementation and one cached
# session, regardless of which order it imports mcp_pptbuilder vs report_lib.
PPT_API = "https://csresearchhub.com/api/generate-ppt"

_ppt_session: Optional[requests.Session] = None


def _find_hidden_inputs(html: str) -> dict[str, str]:
    """Extract name/value pairs from <input type="hidden"> tags, order-agnostic."""
    inputs = {}
    for tag in re.findall(r"<input\b[^>]*>", html, re.IGNORECASE):
        if not re.search(r'type=["\']hidden["\']', tag, re.IGNORECASE):
            continue
        name_m = re.search(r'name=["\']([^"\']+)["\']', tag)
        value_m = re.search(r'value=["\']([^"\']*)["\']', tag)
        if name_m:
            inputs[html_lib.unescape(name_m.group(1))] = html_lib.unescape(value_m.group(1) if value_m else "")
    return inputs


def _submit_cognito_form(
    session: requests.Session, page_resp: requests.Response, field_name: str, field_value: str
) -> requests.Response:
    """Fill in one field (plus the page's hidden inputs) and POST its <form>."""
    text = page_resp.text
    form_m = re.search(r"<form\b[^>]*>", text, re.IGNORECASE)
    if not form_m:
        raise RuntimeError(
            f"Could not find a Cognito form at {page_resp.url} — the ALB/Cognito flow may have changed."
        )
    action_m = re.search(r'action=["\']([^"\']+)["\']', form_m.group(0))
    if not action_m:
        raise RuntimeError(f"Cognito form at {page_resp.url} has no action URL.")
    action_url = urljoin(page_resp.url, html_lib.unescape(action_m.group(1)))

    form_data = _find_hidden_inputs(text)
    form_data[field_name] = field_value
    # Don't raise_for_status: the final redirect of the OAuth dance lands back
    # on PPT_API itself (e.g. a GET against a POST-only route -> 405), which is
    # a sign login *succeeded*, not a request error. Success is judged by
    # whether the ALB session cookie shows up afterward, not by status code.
    return session.post(action_url, data=form_data, timeout=30)


def _login_to_ppt_builder() -> requests.Session:
    """
    Log in to the Cognito hosted UI in front of PPT_API using the
    PPT_BUILDER_LOGIN / PPT_BUILDER_PASSWORD env vars, and return a Session
    carrying the ALB auth cookies.

    Read lazily (not as module constants) so it doesn't matter whether a
    pipeline imports report_lib before or after mcp_pptbuilder — by the time
    this actually runs (a deck is being built), .env has long since loaded.

    Cognito's managed login UI is a two-step flow: the initial /login form only
    collects the username; submitting it redirects to a separate /verifyPassword
    page (new form, new CSRF/hidden fields) that collects the password.
    """
    login = os.environ.get("PPT_BUILDER_LOGIN")
    password = os.environ.get("PPT_BUILDER_PASSWORD")
    if not login or not password:
        raise RuntimeError(
            "PPT_BUILDER_LOGIN / PPT_BUILDER_PASSWORD are not set; cannot log in to PPT Builder."
        )

    session = requests.Session()

    # Hitting the protected API unauthenticated triggers the ALB -> Cognito
    # hosted-UI redirect; requests follows it and lands on the username page.
    username_page = session.get(PPT_API, timeout=30)
    username_page.raise_for_status()

    password_page = _submit_cognito_form(session, username_page, "username", login)
    auth_resp = _submit_cognito_form(session, password_page, "password", password)

    if not any(c.name.startswith("AWSELBAuthSessionCookie") for c in session.cookies):
        raise RuntimeError(
            f"Cognito login did not establish an ALB session cookie (landed on {auth_resp.url}) — "
            "check PPT_BUILDER_LOGIN / PPT_BUILDER_PASSWORD."
        )

    return session


def get_ppt_session(force_relogin: bool = False) -> requests.Session:
    """Return a cached, logged-in Session for PPT_API, logging in as needed."""
    global _ppt_session
    if force_relogin or _ppt_session is None:
        _ppt_session = _login_to_ppt_builder()
    return _ppt_session


# ── MCP tool loading ─────────────────────────────────────────────────────────
def load_tool(module_name: str, attr: str):
    """Import an mcp/ tool module and return a directly-callable function.

    Adds mcp/ to sys.path (so bare `import mcp_serverv3` works WITHOUT shadowing
    the installed `mcp` SDK), then unwraps the fastmcp @mcp.tool() wrapper
    (`getattr(fn, "fn", fn)` — a no-op when it's already a plain function).
    """
    mcp_dir = str(ROOT / "mcp")
    if mcp_dir not in sys.path:
        sys.path.insert(0, mcp_dir)
    mod = importlib.import_module(module_name)
    fn = getattr(mod, attr)
    return getattr(fn, "fn", fn)
